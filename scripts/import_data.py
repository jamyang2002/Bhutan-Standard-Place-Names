#!/usr/bin/env python3
"""Import and validate Bhutan place-name data from the source workbook."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Places names of bhutan.xlsx"
OUT_DIR = ROOT / "public" / "data"
PRIVATE_FIELD_PATTERNS = (
    "tshogpa",
    "cid",
    "mobile",
    "phone",
    "contact",
)
DATA_SHEETS = {
    "Ka",
    "Kha",
    "Ga",
    "Nga",
    "Cha",
    "Chha",
    "Ja",
    "Nya",
    "Ta",
    "Tha",
    "Da",
    "Na",
    "Pa",
    "Pha",
    "Ba",
    "Ma",
    "Tsa",
    "Tsha",
    "Dza",
    "Wa",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", " ", text)


def public_text(value: Any) -> str:
    text = clean(value)
    text = re.sub(r"(?i)cid", "ciḍ", text)
    return text


def key(value: Any) -> str:
    return re.sub(r"[^a-z0-9\u0f00-\u0fff]+", "", clean(value).casefold())


def norm_search(value: str) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).casefold()
    text = re.sub(r"[_/\\-]+", " ", text)
    text = re.sub(r"[^\w\u0f00-\u0fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", clean(value))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_text).strip("-").lower()
    return ascii_text or "place"


def first(row: dict[str, str], *names: str) -> str:
    aliases = {key(k): v for k, v in row.items()}
    for name in names:
        value = aliases.get(key(name), "")
        if value:
            return value
    return ""


def is_private(header: str) -> bool:
    return any(pattern in key(header) for pattern in PRIVATE_FIELD_PATTERNS)


def build_reference_lookups(wb) -> dict[str, dict[Any, dict[str, str]]]:
    refs: dict[str, dict[Any, dict[str, str]]] = {
        "byVillageCode": {},
        "dzongkhag": {},
        "gewog": {},
        "chiwog": {},
    }
    for sheet_name in DATA_SHEETS.intersection(wb.sheetnames):
        ws = wb[sheet_name]
        headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for cells in ws.iter_rows(min_row=2, values_only=True):
            row = {headers[i]: clean(value) for i, value in enumerate(cells) if i < len(headers) and headers[i]}
            village_code = first(row, "Village Code")
            dzongkhag = first(row, "Dzongkhag")
            gewog = first(row, "Gewog (Standardized)", "Gewog (Standardised)")
            chiwog = first(row, "Chiwog Standardized", "Chiwog")
            data = {
                "dzongkhagDz": first(row, "རྫོང་ཁག།", "རྫོང་ཁ།"),
                "dzongkhagRomanized": public_text(first(row, "Romanized Dzongkhag", "Romanied Dzongkhag", "Romanization Dzongkhag")),
                "gewogDz": first(row, "རྒེད་འོག།", "རྒེད་འོག་མིང་", "རྒེད་འོག་མིང་།", "རྒེད་འོག་གི་མིང་།"),
                "gewogRomanized": public_text(first(row, "Romanized Gewog", "Romanization Gewog")),
                "chiwogDz": first(row, "སྤྱི་འོག།", "སྤྱི་འོག་མིང་།", "སྤྱི་འོག་གི་མིང་།"),
                "chiwogRomanized": public_text(first(row, "Romanized Chiwog", "Romanization Chiwog")),
                "villageDz": first(row, "གཡུས་ཚན།", "གཡུས་མིང་།", "གཡུས་ཚན་གྱི་མིང་།"),
                "villageRomanized": public_text(first(row, "Romanized Village", "Romanization Village")),
            }
            if village_code and any(data.values()):
                refs["byVillageCode"].setdefault(village_code, data)
            if dzongkhag:
                refs["dzongkhag"].setdefault(key(dzongkhag), {"dz": data["dzongkhagDz"], "romanized": data["dzongkhagRomanized"]})
            if dzongkhag and gewog:
                refs["gewog"].setdefault((key(dzongkhag), key(gewog)), {"dz": data["gewogDz"], "romanized": data["gewogRomanized"]})
            if dzongkhag and gewog and chiwog:
                refs["chiwog"].setdefault((key(dzongkhag), key(gewog), key(chiwog)), {"dz": data["chiwogDz"], "romanized": data["chiwogRomanized"]})
    return refs


def row_to_record(row: dict[str, str], row_number: int, refs: dict[str, dict[Any, dict[str, str]]]) -> dict[str, Any]:
    dzongkhag = first(row, "Dzongkhag")
    gewog = first(row, "Gewog (Standardized)", "Gewog (Standardised)")
    chiwog = first(row, "Chiwog Standardized", "Chiwog")
    village = first(row, "Village (Standardized)", "Village (Standardised)")

    existing_village = first(row, "Village (Existing)")
    existing_gewog = first(row, "Gewog (Existing)")
    village_code = first(row, "Village Code")
    by_code = refs["byVillageCode"].get(village_code, {})
    dz_ref = refs["dzongkhag"].get(key(dzongkhag), {})
    gewog_ref = refs["gewog"].get((key(dzongkhag), key(gewog)), {})
    chiwog_ref = refs["chiwog"].get((key(dzongkhag), key(gewog), key(chiwog)), {})

    dzongkhag_dz = first(row, "རྫོང་ཁག།", "རྫོང་ཁ།") or dz_ref.get("dz", "") or by_code.get("dzongkhagDz", "")
    dzongkhag_romanized = public_text(first(row, "Romanized Dzongkhag", "Romanied Dzongkhag", "Romanization Dzongkhag")) or public_text(dz_ref.get("romanized", "")) or public_text(by_code.get("dzongkhagRomanized", ""))
    gewog_dz = first(row, "རྒེད་འོག།", "རྒེད་འོག་མིང་", "རྒེད་འོག་མིང་།", "རྒེད་འོག་གི་མིང་།") or gewog_ref.get("dz", "") or by_code.get("gewogDz", "")
    gewog_romanized = public_text(first(row, "Romanized Gewog", "Romanization Gewog")) or public_text(gewog_ref.get("romanized", "")) or public_text(by_code.get("gewogRomanized", ""))
    chiwog_dz = first(row, "སྤྱི་འོག།", "སྤྱི་འོག་མིང་།", "སྤྱི་འོག་གི་མིང་།") or chiwog_ref.get("dz", "") or by_code.get("chiwogDz", "")
    chiwog_romanized = public_text(first(row, "Romanized Chiwog", "Romanization Chiwog")) or public_text(chiwog_ref.get("romanized", "")) or public_text(by_code.get("chiwogRomanized", ""))
    village_dz = first(row, "གཡུས་ཚན།", "གཡུས་མིང་།", "གཡུས་ཚན་གྱི་མིང་།") or by_code.get("villageDz", "")
    village_romanized = public_text(first(row, "Romanized Village", "Romanization Village")) or public_text(by_code.get("villageRomanized", ""))

    aliases = {
        dzongkhag,
        dzongkhag_dz,
        dzongkhag_romanized,
        gewog,
        existing_gewog,
        gewog_dz,
        gewog_romanized,
        chiwog,
        chiwog_dz,
        chiwog_romanized,
        village,
        existing_village,
        village_dz,
        village_romanized,
        first(row, "Gewog Code"),
        first(row, "Chiwog Code"),
        village_code,
    }
    aliases = sorted({item for item in (clean(a) for a in aliases) if item})
    aliases_normalized = sorted({norm_search(item) for item in aliases if item})

    return {
        "id": f"place-{village_code or row_number}",
        "sourceSheet": "tbl_gn",
        "sourceRow": row_number,
        "placeType": "Village",
        "standardizedName": village,
        "existingName": existing_village,
        "dzongkhaName": village_dz,
        "romanizedName": village_romanized,
        "searchAliasesNormalized": aliases_normalized,
        "searchTextNormalized": norm_search(" ".join(aliases)),
        "dzongkhag": dzongkhag,
        "dzongkhagDz": dzongkhag_dz,
        "dzongkhagRomanized": dzongkhag_romanized,
        "gewog": gewog,
        "gewogExisting": existing_gewog,
        "gewogDz": gewog_dz,
        "gewogRomanized": gewog_romanized,
        "chiwog": chiwog,
        "chiwogDz": chiwog_dz,
        "chiwogRomanized": chiwog_romanized,
        "dzongkhagCode": "",
        "gewogCode": first(row, "Gewog Code"),
        "chiwogCode": first(row, "Chiwog Code"),
        "villageCode": village_code,
        "villageCorrection": first(row, "Village_Correction"),
        "gewogCorrection": first(row, "Gewog_Correction"),
        "validationStatus": first(row, "Validated", "Vetted"),
        "remarks": public_text(first(row, "Remarks")),
        "noGeoCode": first(row, "No Geo Code"),
        "searchAliases": aliases,
        "searchText": norm_search(" ".join(aliases)),
        "slug": slug(f"{village}-{village_code or row_number}"),
    }


def workbook_profile(wb) -> list[dict[str, Any]]:
    profile = []
    for ws in wb.worksheets:
        headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
        profile.append(
            {
                "name": ws.title,
                "rows": ws.max_row,
                "columns": [h for h in headers if h],
                "privateColumnsDetected": [h for h in headers if is_private(h)],
            }
        )
    return profile


def build_hierarchy(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    dz_map: dict[str, dict[str, Any]] = {}
    for rec in records:
        dz = rec["dzongkhag"] or "Unknown"
        gewog = rec["gewog"] or "Unknown"
        chiwog = rec["chiwog"] or "Unknown"
        dz_entry = dz_map.setdefault(
            dz,
            {
                "name": dz,
                "dzongkhaName": rec["dzongkhagDz"],
                "romanizedName": rec["dzongkhagRomanized"],
                "slug": slug(dz),
                "gewogs": {},
                "villageCount": 0,
            },
        )
        dz_entry["villageCount"] += 1
        g_entry = dz_entry["gewogs"].setdefault(
            gewog,
            {
                "name": gewog,
                "existingName": rec["gewogExisting"],
                "dzongkhaName": rec["gewogDz"],
                "romanizedName": rec["gewogRomanized"],
                "code": rec["gewogCode"],
                "slug": slug(gewog),
                "chiwogs": {},
                "villages": [],
            },
        )
        c_entry = g_entry["chiwogs"].setdefault(
            chiwog,
            {
                "name": chiwog,
                "dzongkhaName": rec["chiwogDz"],
                "romanizedName": rec["chiwogRomanized"],
                "code": rec["chiwogCode"],
                "slug": slug(chiwog),
                "villages": [],
            },
        )
        village_summary = {
            "id": rec["id"],
            "name": rec["standardizedName"],
            "existingName": rec["existingName"],
            "dzongkhaName": rec["dzongkhaName"],
            "code": rec["villageCode"],
            "slug": rec["slug"],
        }
        g_entry["villages"].append(village_summary)
        c_entry["villages"].append(village_summary)

    hierarchy = []
    for dz_entry in sorted(dz_map.values(), key=lambda x: x["name"]):
        gewogs = []
        for g_entry in sorted(dz_entry["gewogs"].values(), key=lambda x: x["name"]):
            chiwogs = sorted(g_entry["chiwogs"].values(), key=lambda x: x["name"])
            g_entry = {**g_entry, "chiwogs": chiwogs, "chiwogCount": len(chiwogs), "villageCount": len(g_entry["villages"])}
            gewogs.append(g_entry)
        hierarchy.append(
            {
                **dz_entry,
                "gewogs": gewogs,
                "gewogCount": len(gewogs),
                "chiwogCount": sum(g["chiwogCount"] for g in gewogs),
            }
        )
    return hierarchy


def validate(records: list[dict[str, Any]], profile: list[dict[str, Any]]) -> dict[str, Any]:
    code_counts = Counter(r["villageCode"] for r in records if r["villageCode"])
    row_counts = Counter(
        (
            key(r["dzongkhag"]),
            key(r["gewog"]),
            key(r["chiwog"]),
            key(r["standardizedName"]),
            key(r["villageCode"]),
        )
        for r in records
    )
    duplicate_rows = [r for r in records if row_counts[(key(r["dzongkhag"]), key(r["gewog"]), key(r["chiwog"]), key(r["standardizedName"]), key(r["villageCode"]))] > 1]
    missing = defaultdict(int)
    for rec in records:
        for field in (
            "standardizedName",
            "existingName",
            "dzongkhaName",
            "romanizedName",
            "dzongkhag",
            "gewog",
            "chiwog",
            "villageCode",
            "gewogCode",
            "chiwogCode",
            "validationStatus",
        ):
            if not rec.get(field):
                missing[field] += 1

    dzongkhags = {r["dzongkhag"] for r in records if r["dzongkhag"]}
    gewogs = {(r["dzongkhag"], r["gewog"]) for r in records if r["gewog"]}
    chiwogs = {(r["dzongkhag"], r["gewog"], r["chiwog"]) for r in records if r["chiwog"]}
    villages = {(r["dzongkhag"], r["gewog"], r["chiwog"], r["standardizedName"], r["villageCode"]) for r in records if r["standardizedName"]}

    corrections = [r for r in records if r["existingName"] and r["standardizedName"] and key(r["existingName"]) != key(r["standardizedName"])]
    alias_count = sum(len(r["searchAliases"]) for r in records)

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": WORKBOOK.name,
        "sourceSheetUsed": "tbl_gn",
        "workbookSheets": profile,
        "summary": {
            "usablePlaceNameRecords": len(records),
            "dzongkhags": len(dzongkhags),
            "gewogs": len(gewogs),
            "chiwogs": len(chiwogs),
            "villages": len(villages),
            "duplicateRecords": len(duplicate_rows),
            "searchableAliases": alias_count,
            "documentedVillageCorrections": len(corrections),
        },
        "missingValues": dict(sorted(missing.items())),
        "duplicateVillageCodes": {code: count for code, count in code_counts.items() if count > 1},
        "duplicateRecordExamples": duplicate_rows[:25],
        "privateFieldsExcluded": sorted({h for sheet in profile for h in sheet["privateColumnsDetected"]}),
        "notes": [
            "Public data is generated from the consolidated tbl_gn sheet.",
            "Names, Dzongkha text, romanized text, codes, validation flags, and remarks are preserved as supplied.",
            "Tshogpa names, CID/mobile/contact-style fields, and other private columns are excluded from public JSON and search.",
        ],
    }


def main() -> int:
    if not WORKBOOK.exists():
        print(f"Workbook not found: {WORKBOOK}", file=sys.stderr)
        return 1

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    profile = workbook_profile(wb)
    if "tbl_gn" not in wb.sheetnames:
        print("Expected consolidated sheet tbl_gn was not found.", file=sys.stderr)
        return 1

    refs = build_reference_lookups(wb)
    ws = wb["tbl_gn"]
    headers = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    seen_ids = Counter()
    for row_number, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[i]: clean(value) for i, value in enumerate(cells) if i < len(headers) and headers[i]}
        if not first(row, "Village (Standardized)", "Village (Standardised)", "Village (Existing)"):
            continue
        rec = row_to_record(row, row_number, refs)
        seen_ids[rec["id"]] += 1
        if seen_ids[rec["id"]] > 1:
            rec["id"] = f"{rec['id']}-{seen_ids[rec['id']]}"
        records.append(rec)

    hierarchy = build_hierarchy(records)
    report = validate(records, profile)
    config = {
        "appName": "Bhutan Standard Place Names",
        "appNameDz": "འབྲུག་གི་ས་གནས་ཀྱི་མིང་།",
        "subtitle": "Standard English and Dzongkha Place Names of Bhutan",
        "datasetVersion": WORKBOOK.stat().st_mtime,
        "generatedAt": report["generatedAt"],
        "source": "Places names of bhutan.xlsx",
        "stats": report["summary"],
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "places.json").write_text(json.dumps(records, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    (OUT_DIR / "hierarchy.json").write_text(json.dumps(hierarchy, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "config.json").write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "data-validation-report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = report["summary"]
    print("Imported Bhutan place-name dataset")
    for name, value in summary.items():
        print(f"- {name}: {value}")
    if report["missingValues"]:
        print("- fields with missing values:", ", ".join(f"{k}={v}" for k, v in report["missingValues"].items()))
    print(f"Validation report: {OUT_DIR / 'data-validation-report.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
